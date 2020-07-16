

# 接口自动化步骤
# 1.excl测试用例准备ok，代码自动读取测试数据
# 2.发送接口请求，得到响应信息
# 3.断言：实际结果和预期结果比较——通过/不通过
# 4.写入通过/不通过到excl里


# 第三方库：操作excl表格————openpyxl库：实现excl读取测试数据，并写入数据
# 1.安装 pip install openpyxl 2.导入
# 注意：把文件拉到pycharm里，lesson6同级，方便读取

# EXcel中三大对象：
    # 1.工作簿workbook
    # 2.表单shell         sheet = wb['register']获取表单
    # 3.单元格cell         cell = sheet.cell(row = 2 , column = 1)通过表单获取行号列号--单元格
    #                      cell = sheet.cell(row = 2 , column = 1).vale获取单元格内元素



import openpyxl
import requests
#读取测试用例函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  #加载工作簿--文档名字
    sheet = wb[sheetname]
    max_row = sheet.max_row      #获取最大行数，把range里尾换成max_row+1
    case_list = []              #创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id =  sheet.cell(row = i , column = 1).value,
        url = sheet.cell(row = i , column = 5).value,  #获取url
        data = sheet.cell(row = i ,column = 6).value,  #获取data
        expect = sheet.cell(row=i, column=7).value
        )
        case_list.append(dict1)     #每循环一次，就把读取到的字典数据存放到list里
    return case_list            #返回测试用例列表


# cases = read_data('test_case_api.xlsx','register')
# print(cases)


#写入结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row = row ,column = column).value = final_result    #写入结果
    wb.save('test_case_api.xlsx')                   #保存文档

# write_result('test_case_api.xlsx','login',3,8,"Failed")


#执行接口函数
def api_fun(url,data):
    headers_log = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url,json=data,headers=headers_log)
    response = res.json()
    return response



# 断言并写回执行结果
cases = read_data('test_case_api.xlsx','register')         #读取数据
for case in cases:
    case_id =  case.get('case_id')                           #或case['case_id']
    url = case.get('url')
    data = eval(case.get('data'))
    expect = eval(case.get('expect'))                      #获取预期结果
    expect_msg = expect.get('msg')                         #获取预期结果中的msg
    real_result = api_fun(url = url,data= data)           #调用接口函数,返回结果用real_result接收。但是现在的数据类型是字符串,
                                                          #引入eval函数，能去掉引号，去除引号内元素
    real_msg = real_result.get('msg')                     #获取实际结果中的msg
    print('预期结果中的msg：{}'.format(expect_msg))
    print('实际结果中的msg：{}'.format(real_msg))
    if real_msg == expect_msg:
        print('第{}条用例执行通过！'.format(case_id))
        final_re = "Passed"
    else:
        print('第{}条用例测试不通过！'.format(case_id))
        final_re = "Failed"
    write_result('test_case_api.xlsx','register',case_id+1,8,final_re)
    print('*'*20)




# 简历写熟悉Python语言，可利用requests及openpyxl库编写接口自动化脚本实现接口自动化测试，不会写自动化框架，但是原来公司自动化框架已经搭建好了，
# 我会往框架里加写自动化脚本。


